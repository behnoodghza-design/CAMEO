#!/usr/bin/env python3
"""
Enterprise Excel Compliance Report Generator — v3 (3-Sheet Unified)
=====================================================================
Generates a LOCKED, enterprise-grade, 3-sheet Excel workbook:

  Sheet 1: "Inventory & EU Compliance"
      #, Name, CAS, EC Number, Quantity, EU H-Codes, SVHC Status,
      Restrictions, Primary Hazard Group

  Sheet 2: "Compatibility Matrix"
      Lower-triangular NxN visual matrix with chemical names on both
      axes. Color-coded cells: Green/Yellow/Red.

  Sheet 3: "Pairwise Analysis"
      Chem A, Chem B, Status, Hazards (verbatim), Gases (verbatim)
      Only includes non-Compatible pairs.

EC FALLBACK LOGIC:
    If CAS lookup returns 0 rows in any EU table, resolve EC number
    from eu_clp_hazards, then re-query by ec_number.

ZERO-LIABILITY RULE:
    All data is queried VERBATIM from the database.
    Missing EU data → exact string: "Not explicitly listed in EU harmonized lists"

SECURITY:
    ALL cells are LOCKED. Password: 'safeware'.
    Users can sort/filter but CANNOT edit.
"""

import os
import sqlite3
import logging
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel_generator")

# ── Constants ────────────────────────────────────────────
NOT_LISTED = "Not explicitly listed in EU harmonized lists"

# ── Corporate Elite Palette ──────────────────────────────
BRAND_BLUE = "1E3A8A"
HEADER_BG = PatternFill(start_color=BRAND_BLUE, end_color=BRAND_BLUE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
SUB_FONT = Font(name="Calibri", color="93C5FD", size=10)

COL_HDR_BG = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
COL_HDR_FONT = Font(name="Calibri", bold=True, color="E0E7FF", size=11)

DATA_FONT = Font(name="Calibri", size=10, color="1E293B")
HAZARD_FONT = Font(name="Calibri", size=10, color="991B1B", bold=True)
SAFE_FONT = Font(name="Calibri", size=10, color="166534", italic=True)
CAUTION_FONT = Font(name="Calibri", size=10, color="854D0E")

# Matrix cell fills (NOAA standard 3 colors)
FILL_COMPATIBLE = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
FILL_CAUTION = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
FILL_INCOMPATIBLE = PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid")
FILL_DIAGONAL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
FILL_EMPTY_UPPER = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

# Matrix cell fonts
MATRIX_COMPAT_FONT = Font(name="Calibri", size=8, color="166534")
MATRIX_CAUTION_FONT = Font(name="Calibri", size=8, color="854D0E", bold=True)
MATRIX_INCOMPAT_FONT = Font(name="Calibri", size=8, color="991B1B", bold=True)
MATRIX_LABEL_FONT = Font(name="Calibri", size=9, color="334155", bold=True)

# Shared styles
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LOCKED = Protection(locked=True)


# ═══════════════════════════════════════════════════════════
#  Compliance Data Layer (with EC Fallback)
# ═══════════════════════════════════════════════════════════

def query_eu_compliance(db_path: str, cas_numbers: list[str]) -> list[dict]:
    """
    Query all EU compliance data for given CAS numbers.
    Uses EC fallback: if CAS returns 0 rows, resolve EC from eu_clp_hazards
    and re-query by ec_number.

    Returns a list of dicts with keys:
        cas_number, ec_number, chemical_name, cameo_hazards,
        eu_hcodes, classification, svhc_status, restrictions
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    results = []
    for cas in cas_numbers:
        cas_clean = cas.strip()

        # ── 1. CAMEO data ──
        cur.execute("""
            SELECT c.name, c.special_hazards
            FROM chemicals c
            INNER JOIN chemical_cas cc ON c.id = cc.chem_id
            WHERE cc.cas_id = ?
            LIMIT 1
        """, (cas_clean,))
        cameo = cur.fetchone()
        chemical_name = cameo["name"] if cameo else cas_clean
        cameo_hazards = (cameo["special_hazards"] if cameo and cameo["special_hazards"] else NOT_LISTED)

        # ── 2. EU CLP (+ resolve EC number) ──
        cur.execute("""
            SELECT ec_number, classification, hazard_statement_codes, pictogram_signal_codes
            FROM eu_clp_hazards WHERE cas_number = ?
        """, (cas_clean,))
        clp_rows = cur.fetchall()

        ec_number = None
        if clp_rows:
            ec_number = clp_rows[0]["ec_number"]
            hcodes = set()
            classifications = set()
            for row in clp_rows:
                if row["hazard_statement_codes"]:
                    for c in row["hazard_statement_codes"].replace("\n", " ").split():
                        hcodes.add(c.strip())
                if row["classification"]:
                    for cl in row["classification"].split("\n"):
                        cl = cl.strip()
                        if cl:
                            classifications.add(cl)
            eu_hcodes = "\n".join(sorted(hcodes)) if hcodes else NOT_LISTED
            classification = "\n".join(sorted(classifications)) if classifications else ""
        else:
            eu_hcodes = NOT_LISTED
            classification = ""

        # ── 3. SVHC (CAS → EC fallback) ──
        cur.execute("SELECT substance_name, reason_for_inclusion, date_of_inclusion FROM eu_svhc WHERE cas_number = ?", (cas_clean,))
        svhc_rows = cur.fetchall()

        if not svhc_rows and ec_number:
            cur.execute("SELECT substance_name, reason_for_inclusion, date_of_inclusion FROM eu_svhc WHERE ec_number = ?", (ec_number,))
            svhc_rows = cur.fetchall()

        if svhc_rows:
            parts = []
            for row in svhc_rows:
                reason = row["reason_for_inclusion"] or "Listed"
                date = row["date_of_inclusion"] or ""
                parts.append(f"SVHC — {reason}" + (f" ({date})" if date else ""))
            svhc_status = "\n".join(parts)
        else:
            svhc_status = NOT_LISTED

        # ── 4. Restrictions (CAS → EC fallback) ──
        cur.execute("SELECT entry_number, conditions, remarks FROM eu_restrictions WHERE cas_number = ?", (cas_clean,))
        rest_rows = cur.fetchall()

        if not rest_rows and ec_number:
            cur.execute("SELECT entry_number, conditions, remarks FROM eu_restrictions WHERE ec_number = ?", (ec_number,))
            rest_rows = cur.fetchall()

        if rest_rows:
            parts = []
            for row in rest_rows:
                entry = row["entry_number"] or "?"
                remark = row["remarks"] or row["conditions"] or "Restricted"
                parts.append(f"Entry {entry}: {remark}")
            restrictions = "\n".join(parts)
        else:
            restrictions = NOT_LISTED

        results.append({
            "cas_number": cas_clean,
            "ec_number": ec_number or "",
            "chemical_name": chemical_name,
            "cameo_hazards": cameo_hazards,
            "eu_hcodes": eu_hcodes,
            "classification": classification,
            "svhc_status": svhc_status,
            "restrictions": restrictions,
        })

    conn.close()
    return results


# ═══════════════════════════════════════════════════════════
#  Sheet Helpers
# ═══════════════════════════════════════════════════════════

def _apply_sheet_protection(ws):
    """Lock sheet: password 'safeware', allow sort/filter only."""
    ws.protection.sheet = True
    ws.protection.password = "safeware"
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False


def _write_enterprise_header(ws, title: str, subtitle: str, num_cols: int):
    """Write rows 1-3: title, subtitle, spacer."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = HEADER_FONT
    c.fill = HEADER_BG
    c.alignment = CENTER
    c.protection = LOCKED
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = SUB_FONT
    c2.fill = HEADER_BG
    c2.alignment = CENTER
    c2.protection = LOCKED
    ws.row_dimensions[2].height = 22

    ws.row_dimensions[3].height = 5
    for ci in range(1, num_cols + 1):
        ws.cell(row=3, column=ci).fill = HEADER_BG
        ws.cell(row=3, column=ci).protection = LOCKED


def _write_column_headers(ws, row: int, headers: list[tuple[str, int]]):
    """Write column headers with widths."""
    for ci, (text, width) in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=ci, value=text)
        cell.font = COL_HDR_FONT
        cell.fill = COL_HDR_BG
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        cell.protection = LOCKED
        ws.column_dimensions[get_column_letter(ci)].width = width
    return len(headers)


# ═══════════════════════════════════════════════════════════
#  ComplianceExcelGenerator
# ═══════════════════════════════════════════════════════════

class ComplianceExcelGenerator:
    """Enterprise Excel report generator for EU REACH/CLP compliance."""

    def __init__(self, db_path: str):
        if not os.path.exists(db_path):
            raise FileNotFoundError(f"Database not found: {db_path}")
        self.db_path = db_path

    # ── Legacy single-sheet (backward compat) ──
    def generate(
        self,
        cas_numbers: list[str],
        output_path: str,
        report_title: Optional[str] = None,
    ) -> str:
        """Generate single-sheet compliance report (backward compatible)."""
        data = query_eu_compliance(self.db_path, cas_numbers)
        return self._write_single_sheet(data, output_path, report_title)

    # ══════════════════════════════════════════════════════════
    #  UNIFIED 3-SHEET WORKBOOK
    # ══════════════════════════════════════════════════════════

    def generate_unified(
        self,
        inventory_data: list[dict],
        chemicals: list[dict],
        reactivity_pairs: list[dict],
        matrix_data: list[list[dict]],
        output_path: str,
        report_title: Optional[str] = None,
    ) -> str:
        """
        Generate 3-sheet unified enterprise report.

        Args:
            inventory_data: EU compliance records (from query_eu_compliance)
            chemicals: List of {id, name, cas_ids, ...} from analysis payload
            reactivity_pairs: List of pair_details dicts from analysis
            matrix_data: NxN matrix from analysis payload (list of list of dicts)
            output_path: Destination file path
            report_title: Optional custom title
        """
        logger.info("Generating 3-sheet unified report: %d chemicals, %d pairs",
                     len(inventory_data), len(reactivity_pairs))

        wb = Workbook()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # ════════════════════════════════════════════════════════
        # SHEET 1: Inventory & EU Compliance
        # ════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Inventory & EU Compliance"

        s1_title = report_title or "SAFEWARE — INVENTORY & EU COMPLIANCE REPORT"
        s1_sub = f"Generated: {timestamp}  |  Items: {len(inventory_data)}  |  Sources: CAMEO, ECHA CLP, SVHC Candidate List, REACH Annex XVII"

        S1_COLS = [
            ("#", 5),
            ("Chemical Name", 32),
            ("CAS Number", 14),
            ("EC Number", 14),
            ("Quantity", 12),
            ("EU CLP H-Codes", 28),
            ("SVHC Status", 36),
            ("EU Restrictions", 36),
            ("Primary Hazard Group", 28),
        ]
        nc1 = len(S1_COLS)

        _write_enterprise_header(ws1, s1_title, s1_sub, nc1)
        hdr1 = 4
        _write_column_headers(ws1, hdr1, S1_COLS)

        for idx, rec in enumerate(inventory_data):
            r = hdr1 + 1 + idx
            vals = [
                idx + 1,
                rec.get("chemical_name", ""),
                rec.get("cas_number", ""),       # CAS in CAS column (not name)
                rec.get("ec_number", ""),
                rec.get("quantity", ""),
                rec.get("eu_hcodes", NOT_LISTED),
                rec.get("svhc_status", NOT_LISTED),
                rec.get("restrictions", NOT_LISTED),
                rec.get("classification", ""),
            ]

            for ci, val in enumerate(vals, start=1):
                cell = ws1.cell(row=r, column=ci, value=val if val else NOT_LISTED if ci >= 6 else "")
                cell.alignment = WRAP
                cell.border = THIN_BORDER
                cell.protection = LOCKED

                sval = str(val)
                if val == NOT_LISTED:
                    cell.font = SAFE_FONT
                elif ci >= 6 and sval != NOT_LISTED and sval:
                    cell.font = HAZARD_FONT
                    if "SVHC" in sval:
                        cell.fill = FILL_INCOMPATIBLE
                else:
                    cell.font = DATA_FONT

        last1 = hdr1 + max(len(inventory_data), 1)
        ws1.auto_filter.ref = f"A{hdr1}:{get_column_letter(nc1)}{last1}"
        ws1.freeze_panes = f"A{hdr1 + 1}"
        ws1.page_setup.orientation = "landscape"
        ws1.page_setup.fitToWidth = 1
        ws1.print_title_rows = f"1:{hdr1}"
        _apply_sheet_protection(ws1)

        # ════════════════════════════════════════════════════════
        # SHEET 2: Compatibility Matrix (Lower-Triangular)
        # ════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Compatibility Matrix")

        chem_names = [c.get("name", f"Chemical {i+1}") for i, c in enumerate(chemicals)]
        n = len(chem_names)

        s2_title = "CHEMICAL COMPATIBILITY MATRIX"
        s2_sub = f"Generated: {timestamp}  |  Chemicals: {n}  |  Lower-triangular (symmetric)"
        total_cols_2 = n + 1  # label column + n chemicals

        _write_enterprise_header(ws2, s2_title, s2_sub, max(total_cols_2, 3))

        # Column headers (row 4): blank + chemical names
        hdr2 = 4
        ws2.column_dimensions["A"].width = 28
        cell_corner = ws2.cell(row=hdr2, column=1, value="")
        cell_corner.fill = COL_HDR_BG
        cell_corner.border = THIN_BORDER
        cell_corner.protection = LOCKED

        for ci, name in enumerate(chem_names):
            col = ci + 2
            cell = ws2.cell(row=hdr2, column=col, value=name)
            cell.font = MATRIX_LABEL_FONT
            cell.fill = COL_HDR_BG
            cell.alignment = Alignment(text_rotation=45, horizontal="center", vertical="bottom")
            cell.border = THIN_BORDER
            cell.protection = LOCKED
            ws2.column_dimensions[get_column_letter(col)].width = 5

        # Matrix body (rows 5+)
        for i in range(n):
            r = hdr2 + 1 + i
            # Row label
            lbl = ws2.cell(row=r, column=1, value=chem_names[i])
            lbl.font = MATRIX_LABEL_FONT
            lbl.alignment = Alignment(horizontal="right", vertical="center")
            lbl.border = THIN_BORDER
            lbl.protection = LOCKED

            for j in range(n):
                col = j + 2
                cell = ws2.cell(row=r, column=col)
                cell.border = THIN_BORDER
                cell.protection = LOCKED
                cell.alignment = CENTER

                if i == j:
                    # Diagonal
                    cell.value = "—"
                    cell.fill = FILL_DIAGONAL
                    cell.font = Font(name="Calibri", size=9, color="94A3B8")
                elif j > i:
                    # Upper triangle → empty
                    cell.fill = FILL_EMPTY_UPPER
                else:
                    # Lower triangle → actual compatibility
                    # matrix_data[i][j] should mirror matrix_data[j][i]
                    mat_cell = None
                    if matrix_data and i < len(matrix_data) and j < len(matrix_data[i]):
                        mat_cell = matrix_data[i][j]

                    status = (mat_cell or {}).get("status", "Compatible")
                    code = {"Compatible": "C", "Caution": "C!", "Incompatible": "X"}.get(status, "C")
                    cell.value = code

                    if status == "Incompatible":
                        cell.fill = FILL_INCOMPATIBLE
                        cell.font = MATRIX_INCOMPAT_FONT
                    elif status == "Caution":
                        cell.fill = FILL_CAUTION
                        cell.font = MATRIX_CAUTION_FONT
                    else:
                        cell.fill = FILL_COMPATIBLE
                        cell.font = MATRIX_COMPAT_FONT

            ws2.row_dimensions[r].height = 18

        # Legend row
        legend_row = hdr2 + n + 2
        legends = [
            ("C = Compatible", FILL_COMPATIBLE),
            ("C! = Caution", FILL_CAUTION),
            ("X = Incompatible", FILL_INCOMPATIBLE),
        ]
        for li, (text, fill) in enumerate(legends):
            lc = ws2.cell(row=legend_row, column=li + 1, value=text)
            lc.fill = fill
            lc.font = DATA_FONT
            lc.border = THIN_BORDER
            lc.protection = LOCKED

        ws2.freeze_panes = f"B{hdr2 + 1}"
        ws2.page_setup.orientation = "landscape"
        ws2.page_setup.fitToWidth = 1
        _apply_sheet_protection(ws2)

        # ════════════════════════════════════════════════════════
        # SHEET 3: Pairwise Analysis (Non-Compatible Only)
        # ════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Pairwise Analysis")

        non_compat = [p for p in reactivity_pairs if p.get("status", "") != "Compatible"]

        s3_title = "DETAILED PAIRWISE ANALYSIS"
        s3_sub = f"Generated: {timestamp}  |  Non-Compatible Pairs: {len(non_compat)}  |  Data Source: CAMEO Reactivity Database"

        S3_COLS = [
            ("#", 5),
            ("Chemical A", 28),
            ("Chemical B", 28),
            ("Compatibility", 16),
            ("Hazards (verbatim)", 45),
            ("Predicted Gases (verbatim)", 40),
            ("Notes", 35),
        ]
        nc3 = len(S3_COLS)

        _write_enterprise_header(ws3, s3_title, s3_sub, nc3)
        hdr3 = 4
        _write_column_headers(ws3, hdr3, S3_COLS)

        for idx, pair in enumerate(non_compat):
            r = hdr3 + 1 + idx
            status = pair.get("status", "Caution")
            hazards = pair.get("hazards", [])
            gases = pair.get("gases", [])
            notes = pair.get("explanation", "")

            vals = [
                idx + 1,
                pair.get("chemical_a_name", ""),
                pair.get("chemical_b_name", ""),
                status,
                ", ".join(hazards) if hazards else "—",
                ", ".join(gases) if gases else "—",
                notes,
            ]

            row_fill = FILL_INCOMPATIBLE if status == "Incompatible" else FILL_CAUTION
            row_font = HAZARD_FONT if status == "Incompatible" else CAUTION_FONT

            for ci, val in enumerate(vals, start=1):
                cell = ws3.cell(row=r, column=ci, value=val)
                cell.alignment = WRAP
                cell.border = THIN_BORDER
                cell.protection = LOCKED
                cell.fill = row_fill
                cell.font = row_font if ci >= 4 else DATA_FONT

        if not non_compat:
            r = hdr3 + 1
            ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=nc3)
            cell = ws3.cell(row=r, column=1, value="All chemicals in this inventory are compatible. No hazardous pairs detected.")
            cell.font = SAFE_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            cell.protection = LOCKED

        last3 = hdr3 + max(len(non_compat), 1)
        ws3.auto_filter.ref = f"A{hdr3}:{get_column_letter(nc3)}{last3}"
        ws3.freeze_panes = f"A{hdr3 + 1}"
        ws3.page_setup.orientation = "landscape"
        ws3.page_setup.fitToWidth = 1
        ws3.print_title_rows = f"1:{hdr3}"
        _apply_sheet_protection(ws3)

        # ════ Save ════
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info("3-sheet unified report saved: %s", output_path)
        return os.path.abspath(output_path)

    # ── Legacy single-sheet writer ──
    def _write_single_sheet(self, data: list[dict], output_path: str, title: Optional[str]) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Compliance Report"
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        COLS = [
            ("CAS Number", 18),
            ("EC Number", 14),
            ("Chemical Name", 35),
            ("CAMEO Hazards", 40),
            ("EU CLP H-Codes", 30),
            ("SVHC Status", 40),
            ("EU Restrictions", 40),
        ]
        nc = len(COLS)

        _write_enterprise_header(
            ws,
            title or "SAFEWARE ENTERPRISE COMPLIANCE REPORT",
            f"Generated: {timestamp}  |  Items: {len(data)}  |  Sources: CAMEO, ECHA CLP, SVHC, Annex XVII",
            nc,
        )
        hdr = 4
        _write_column_headers(ws, hdr, COLS)

        for idx, rec in enumerate(data):
            r = hdr + 1 + idx
            vals = [
                rec["cas_number"], rec.get("ec_number", ""),
                rec["chemical_name"], rec["cameo_hazards"],
                rec["eu_hcodes"], rec["svhc_status"], rec["restrictions"],
            ]
            for ci, val in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=ci, value=val)
                cell.alignment = WRAP
                cell.border = THIN_BORDER
                cell.protection = LOCKED
                sval = str(val)
                if val == NOT_LISTED:
                    cell.font = SAFE_FONT
                elif ci >= 4 and sval != NOT_LISTED and sval:
                    cell.font = HAZARD_FONT
                    if "SVHC" in sval:
                        cell.fill = FILL_INCOMPATIBLE
                else:
                    cell.font = DATA_FONT

        last = hdr + len(data)
        ws.auto_filter.ref = f"A{hdr}:{get_column_letter(nc)}{last}"
        ws.freeze_panes = f"A{hdr + 1}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.print_title_rows = f"1:{hdr}"
        _apply_sheet_protection(ws)

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info("Report saved: %s (%d records)", output_path, len(data))
        return os.path.abspath(output_path)
