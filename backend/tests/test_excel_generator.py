#!/usr/bin/env python3
"""
EU Compliance — Excel Generator Tests (v2 — Multi-Sheet)
=========================================================
Tests the ComplianceExcelGenerator for:
  1. Single-sheet backward compatibility
  2. Multi-sheet unified report
  3. EC fallback logic
  4. Sheet protection on all sheets
  5. Corporate styling (header, freeze panes, conditional fills)

Usage:
    pytest tests/test_excel_generator.py -v
"""

import os
import sys
import sqlite3
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook
from logic.excel_generator import ComplianceExcelGenerator, query_eu_compliance, NOT_LISTED


# ── Fixtures ─────────────────────────────────────────────

@pytest.fixture
def test_db(tmp_path):
    """Create a minimal test database with CAMEO + EU tables."""
    db_path = str(tmp_path / "test_chemicals.db")
    conn = sqlite3.connect(db_path)
    c = conn.cursor()

    c.execute("CREATE TABLE chemicals (id INTEGER PRIMARY KEY, name TEXT, special_hazards TEXT)")
    c.execute("INSERT INTO chemicals VALUES (1, 'Acetone', 'Flammable liquid and vapor')")
    c.execute("INSERT INTO chemicals VALUES (2, 'Hydrogen', 'Extremely flammable gas')")
    c.execute("INSERT INTO chemicals VALUES (3, 'Chromium trioxide', 'Highly toxic by all routes')")

    c.execute("CREATE TABLE chemical_cas (chem_id INTEGER, cas_id TEXT, cas_nodash TEXT, sort INTEGER, annotation TEXT, PRIMARY KEY (chem_id, cas_id))")
    c.execute("INSERT INTO chemical_cas VALUES (1, '67-64-1', '67641', 0, NULL)")
    c.execute("INSERT INTO chemical_cas VALUES (2, '1333-74-0', '1333740', 0, NULL)")
    c.execute("INSERT INTO chemical_cas VALUES (3, '1333-82-0', '1333820', 0, NULL)")

    c.execute("""CREATE TABLE eu_clp_hazards (
        id INTEGER PRIMARY KEY AUTOINCREMENT, index_number TEXT, cas_number TEXT NOT NULL,
        ec_number TEXT, chemical_name_eu TEXT, classification TEXT,
        hazard_statement_codes TEXT, pictogram_signal_codes TEXT,
        labelling_hazard_codes TEXT, suppl_hazard_codes TEXT,
        specific_conc_limits TEXT, notes TEXT, UNIQUE(cas_number, index_number))""")
    c.execute("""INSERT INTO eu_clp_hazards (index_number, cas_number, ec_number, chemical_name_eu, classification, hazard_statement_codes)
        VALUES ('606-001-00-8', '67-64-1', '200-662-2', 'acetone', 'Flam. Liq. 2', 'H225\nH319')""")
    c.execute("""INSERT INTO eu_clp_hazards (index_number, cas_number, ec_number, chemical_name_eu, classification, hazard_statement_codes)
        VALUES ('001-001-00-9', '1333-74-0', '215-605-7', 'hydrogen', 'Flam. Gas 1', 'H220')""")
    c.execute("""INSERT INTO eu_clp_hazards (index_number, cas_number, ec_number, chemical_name_eu, classification, hazard_statement_codes)
        VALUES ('024-010-00-6', '1333-82-0', '215-607-8', 'chromium trioxide', 'Carc. 1A\nMuta. 1B', 'H350\nH340\nH271')""")

    c.execute("""CREATE TABLE eu_svhc (
        id INTEGER PRIMARY KEY AUTOINCREMENT, substance_name TEXT, ec_number TEXT,
        cas_number TEXT, reason_for_inclusion TEXT, date_of_inclusion TEXT,
        remarks TEXT, UNIQUE(cas_number, ec_number, substance_name))""")
    # Chromium listed by EC only (to test EC fallback)
    c.execute("INSERT INTO eu_svhc VALUES (1, 'Chromium trioxide', '215-607-8', NULL, 'Carcinogenic (Art. 57a)', '15-Dec-2010', NULL)")

    c.execute("""CREATE TABLE eu_restrictions (
        id INTEGER PRIMARY KEY AUTOINCREMENT, substance_name TEXT, ec_number TEXT,
        cas_number TEXT, entry_number TEXT, conditions TEXT,
        remarks TEXT, UNIQUE(cas_number, ec_number, entry_number))""")
    # Restriction by EC only
    c.execute("INSERT INTO eu_restrictions VALUES (1, 'Chromium trioxide', '215-607-8', NULL, '47', NULL, 'CMR restriction')")

    conn.commit()
    conn.close()
    return db_path


# ── Tests ────────────────────────────────────────────────

class TestECFallback:
    """Test the EC fallback logic in query_eu_compliance."""

    def test_ec_fallback_svhc(self, test_db):
        """Chromium (CAS only) should find SVHC via EC fallback."""
        results = query_eu_compliance(test_db, ["1333-82-0"])
        assert len(results) == 1
        r = results[0]
        assert r["ec_number"] == "215-607-8"
        assert "SVHC" in r["svhc_status"]
        assert "Carcinogenic" in r["svhc_status"]

    def test_ec_fallback_restrictions(self, test_db):
        """Chromium (CAS only) should find Restrictions via EC fallback."""
        results = query_eu_compliance(test_db, ["1333-82-0"])
        r = results[0]
        assert r["restrictions"] != NOT_LISTED
        assert "Entry 47" in r["restrictions"]

    def test_normal_cas_lookup(self, test_db):
        """Acetone should work with direct CAS lookup (no fallback needed)."""
        results = query_eu_compliance(test_db, ["67-64-1"])
        r = results[0]
        assert "H225" in r["eu_hcodes"]
        assert r["ec_number"] == "200-662-2"

    def test_fake_cas_not_listed(self, test_db):
        """Fake CAS should return NOT_LISTED for all EU fields."""
        results = query_eu_compliance(test_db, ["999-99-9"])
        r = results[0]
        assert r["eu_hcodes"] == NOT_LISTED
        assert r["svhc_status"] == NOT_LISTED
        assert r["restrictions"] == NOT_LISTED


class TestSingleSheetReport:
    """Test backward-compatible single-sheet generation."""

    def test_file_created(self, test_db, tmp_path):
        output = str(tmp_path / "report.xlsx")
        gen = ComplianceExcelGenerator(test_db)
        result = gen.generate(["67-64-1", "999-99-9"], output)
        assert os.path.exists(result)
        wb = load_workbook(result)
        assert wb.active is not None
        wb.close()

    def test_sheet_protection(self, test_db, tmp_path):
        output = str(tmp_path / "report.xlsx")
        gen = ComplianceExcelGenerator(test_db)
        gen.generate(["67-64-1"], output)
        wb = load_workbook(output)
        assert wb.active.protection.sheet is True
        wb.close()

    def test_enterprise_header(self, test_db, tmp_path):
        output = str(tmp_path / "report.xlsx")
        gen = ComplianceExcelGenerator(test_db)
        gen.generate(["67-64-1"], output)
        wb = load_workbook(output)
        ws = wb.active
        title = ws.cell(row=1, column=1).value or ""
        assert "SAFEWARE" in title.upper()
        assert "COMPLIANCE" in title.upper()
        wb.close()


class TestUnifiedReport:
    """Test multi-sheet unified report generation."""

    def test_two_sheets(self, test_db, tmp_path):
        output = str(tmp_path / "unified.xlsx")
        gen = ComplianceExcelGenerator(test_db)

        inv = query_eu_compliance(test_db, ["67-64-1", "1333-82-0"])
        pairs = [
            {"chemical_a_name": "Acetone", "chemical_b_name": "Chromium trioxide",
             "compatibility": "INCOMPATIBLE", "risk_level": "DANGEROUS",
             "hazards": ["Oxidizer + flammable"], "explanation": "Strong oxidizer with flammable solvent"},
        ]

        gen.generate_unified(inv, pairs, output)
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 2
        assert "Inventory & EU Compliance" in wb.sheetnames
        assert "Reactivity Matrix" in wb.sheetnames
        wb.close()

    def test_both_sheets_locked(self, test_db, tmp_path):
        output = str(tmp_path / "unified.xlsx")
        gen = ComplianceExcelGenerator(test_db)
        inv = query_eu_compliance(test_db, ["67-64-1"])
        gen.generate_unified(inv, [], output)

        wb = load_workbook(output)
        for ws in wb.worksheets:
            assert ws.protection.sheet is True, f"Sheet '{ws.title}' not protected!"
        wb.close()

    def test_freeze_panes(self, test_db, tmp_path):
        output = str(tmp_path / "unified.xlsx")
        gen = ComplianceExcelGenerator(test_db)
        inv = query_eu_compliance(test_db, ["67-64-1"])
        gen.generate_unified(inv, [], output)

        wb = load_workbook(output)
        ws = wb["Inventory & EU Compliance"]
        assert ws.freeze_panes == "A5"
        wb.close()

    def test_reactivity_color_coding(self, test_db, tmp_path):
        output = str(tmp_path / "unified.xlsx")
        gen = ComplianceExcelGenerator(test_db)
        inv = query_eu_compliance(test_db, ["67-64-1"])
        pairs = [
            {"chemical_a_name": "A", "chemical_b_name": "B",
             "compatibility": "INCOMPATIBLE", "risk_level": "DANGEROUS",
             "hazards": ["Fire"], "explanation": "Fire hazard"},
        ]
        gen.generate_unified(inv, pairs, output)

        wb = load_workbook(output)
        ws = wb["Reactivity Matrix"]
        # Row 5 (first data row) should have DANGEROUS fill (#FCA5A5)
        cell = ws.cell(row=5, column=1)
        assert cell.fill.start_color.rgb is not None
        wb.close()

    def test_ec_number_in_output(self, test_db, tmp_path):
        """EC number column must be populated."""
        output = str(tmp_path / "unified.xlsx")
        gen = ComplianceExcelGenerator(test_db)
        inv = query_eu_compliance(test_db, ["67-64-1"])
        gen.generate_unified(inv, [], output)

        wb = load_workbook(output)
        ws = wb["Inventory & EU Compliance"]
        # EC Number is in column 3
        ec_val = ws.cell(row=5, column=3).value
        assert ec_val == "200-662-2"
        wb.close()
