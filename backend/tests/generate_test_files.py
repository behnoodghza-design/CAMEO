"""
generate_test_files.py — Generate comprehensive stress test files for ETL validation.

Queries CAMEO database for real chemicals and creates 20 test files covering:
- Column detection challenges (A1-A5)
- Data quality challenges (B1-B5)
- Matching engine challenges (C1-C5)
- Real-world production scenarios (D1-D5)
"""

import sqlite3
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import os
import random
from datetime import datetime, timedelta

# Path to CAMEO database
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'chemicals.db')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), 'data')

# Ensure output directory exists
os.makedirs(OUTPUT_DIR, exist_ok=True)


def get_db_connection():
    """Get connection to CAMEO database."""
    return sqlite3.connect(DB_PATH)


def query_chemicals(query, params=()):
    """Execute a query and return results as list of dicts."""
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute(query, params)
    results = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return results


def get_petroleum_chemicals(limit=50):
    """Get petroleum/petrochemical products from CAMEO."""
    query = """
        SELECT DISTINCT c.id, c.name, cc.cas_id as cas_number
        FROM chemicals c
        LEFT JOIN chemical_cas cc ON c.id = cc.chem_id
        WHERE c.name LIKE '%benzene%' 
           OR c.name LIKE '%toluene%'
           OR c.name LIKE '%xylene%'
           OR c.name LIKE '%gasoline%'
           OR c.name LIKE '%diesel%'
           OR c.name LIKE '%kerosene%'
           OR c.name LIKE '%naphtha%'
           OR c.name LIKE '%crude%'
           OR c.name LIKE '%petroleum%'
           OR c.name LIKE '%propane%'
           OR c.name LIKE '%butane%'
           OR c.name LIKE '%ethylene%'
           OR c.name LIKE '%propylene%'
        ORDER BY RANDOM()
        LIMIT ?
    """
    return query_chemicals(query, (limit,))


def get_industrial_chemicals(limit=50):
    """Get common industrial chemicals."""
    query = """
        SELECT DISTINCT c.id, c.name, cc.cas_id as cas_number
        FROM chemicals c
        LEFT JOIN chemical_cas cc ON c.id = cc.chem_id
        WHERE c.name IN (
            'SULFURIC ACID', 'HYDROCHLORIC ACID', 'NITRIC ACID',
            'SODIUM HYDROXIDE', 'POTASSIUM HYDROXIDE',
            'ACETONE', 'METHANOL', 'ETHANOL', 'ISOPROPANOL',
            'SODIUM CHLORIDE', 'CALCIUM CARBONATE', 'MAGNESIUM SULFATE',
            'AMMONIA', 'CHLORINE', 'HYDROGEN PEROXIDE'
        )
        OR c.name LIKE '%acid%'
        OR c.name LIKE '%hydroxide%'
        OR c.name LIKE '%chloride%'
        OR c.name LIKE '%sulfate%'
        ORDER BY RANDOM()
        LIMIT ?
    """
    return query_chemicals(query, (limit,))


def get_chemicals_with_un(limit=30):
    """Get chemicals with UN numbers (fallback to regular chemicals if UN table missing)."""
    # Try to get chemicals with UN codes, but fall back if table doesn't exist
    try:
        query = """
            SELECT DISTINCT c.id, c.name, cc.cas_id as cas_number, cu.un_code
            FROM chemicals c
            LEFT JOIN chemical_cas cc ON c.id = cc.chem_id
            LEFT JOIN chemical_un cu ON c.id = cu.chem_id
            WHERE cu.un_code IS NOT NULL
            ORDER BY RANDOM()
            LIMIT ?
        """
        return query_chemicals(query, (limit,))
    except:
        # Fallback: just get regular industrial chemicals
        return get_industrial_chemicals(limit)


def get_chemicals_with_formula(limit=30):
    """Get chemicals with formulas."""
    query = """
        SELECT DISTINCT c.id, c.name, cc.cas_id as cas_number, c.formulas
        FROM chemicals c
        LEFT JOIN chemical_cas cc ON c.id = cc.chem_id
        WHERE c.formulas IS NOT NULL AND c.formulas != '[]'
        ORDER BY RANDOM()
        LIMIT ?
    """
    results = query_chemicals(query, (limit,))
    # Parse formulas JSON
    import json
    for r in results:
        try:
            formulas = json.loads(r['formulas'])
            r['formula'] = formulas[0] if formulas else None
        except:
            r['formula'] = None
    return results


def corrupt_cas_to_date(cas):
    """Simulate Excel date corruption of CAS numbers."""
    # CAS like 2001-02-3 becomes date 2001-02-03
    parts = cas.split('-')
    if len(parts) == 3:
        try:
            year = int(parts[0])
            month = int(parts[1])
            day = int(parts[2])
            if 1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                return f"{year:04d}-{month:02d}-{day:02d}"
        except:
            pass
    return cas


# ═══════════════════════════════════════════════════════
#  Category A: Column Detection Challenges
# ═══════════════════════════════════════════════════════

def generate_A1_no_headers():
    """A1: File with no header row."""
    print("Generating A1_NO_HEADERS.xlsx...")
    chems = get_industrial_chemicals(20)
    
    data = []
    for c in chems:
        data.append([
            c['cas_number'] or '',
            c['name'],
            f"{random.randint(10, 500)} L",
        ])
    
    df = pd.DataFrame(data)  # No column names
    filepath = os.path.join(OUTPUT_DIR, 'A1_NO_HEADERS.xlsx')
    df.to_excel(filepath, index=False, header=False)
    print(f"  ✓ Created {filepath}")


def generate_A2_foreign_headers():
    """A2: Headers in multiple foreign languages."""
    print("Generating A2_FOREIGN_HEADERS.xlsx...")
    chems = get_industrial_chemicals(20)
    
    # Headers in Persian, German, French, Spanish, Chinese
    headers = ['نام ماده', 'شماره کس', 'مقدار', 'واحد', 'تامین کننده']
    
    data = []
    for c in chems:
        data.append({
            headers[0]: c['name'],
            headers[1]: c['cas_number'] or '',
            headers[2]: random.randint(10, 500),
            headers[3]: random.choice(['L', 'kg', 'gal']),
            headers[4]: random.choice(['Supplier A', 'Supplier B', 'Supplier C']),
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'A2_FOREIGN_HEADERS.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_A3_abbreviated_headers():
    """A3: Heavily abbreviated headers."""
    print("Generating A3_ABBREVIATED_HEADERS.xlsx...")
    chems = get_industrial_chemicals(20)
    
    data = []
    for c in chems:
        data.append({
            'chm': c['name'],
            'cas': c['cas_number'] or '',
            'qty': random.randint(10, 500),
            'unt': random.choice(['L', 'kg', 'gal']),
            'loc': random.choice(['A1', 'B2', 'C3']),
            'supp': random.choice(['SupplierX', 'SupplierY']),
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'A3_ABBREVIATED_HEADERS.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_A4_swapped_columns():
    """A4: CAS and Name columns swapped."""
    print("Generating A4_SWAPPED_COLUMNS.xlsx...")
    chems = get_industrial_chemicals(20)
    
    data = []
    for c in chems:
        # Swap: CAS in first column, Name in second
        data.append({
            'Chemical ID': c['cas_number'] or '',  # Actually CAS
            'CAS Number': c['name'],  # Actually name
            'Quantity': f"{random.randint(10, 500)} L",
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'A4_SWAPPED_COLUMNS.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_A5_merged_headers():
    """A5: Excel file with merged cells and multi-row headers."""
    print("Generating A5_MERGED_HEADERS.xlsx...")
    chems = get_industrial_chemicals(15)
    
    filepath = os.path.join(OUTPUT_DIR, 'A5_MERGED_HEADERS.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Row 1: Category headers (merged)
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Chemical Information'
    ws.merge_cells('C1:D1')
    ws['C1'] = 'Inventory Details'
    
    # Row 2: Actual field names
    ws['A2'] = 'Chemical Name'
    ws['B2'] = 'CAS No.'
    ws['C2'] = 'Quantity'
    ws['D2'] = 'Unit'
    
    # Row 3+: Data
    for i, c in enumerate(chems, start=3):
        ws[f'A{i}'] = c['name']
        ws[f'B{i}'] = c['cas_number'] or ''
        ws[f'C{i}'] = random.randint(10, 500)
        ws[f'D{i}'] = random.choice(['L', 'kg', 'gal'])
    
    wb.save(filepath)
    print(f"  ✓ Created {filepath}")


# ═══════════════════════════════════════════════════════
#  Category B: Data Quality Challenges
# ═══════════════════════════════════════════════════════

def generate_B1_excel_date_corruption():
    """B1: CAS numbers corrupted to dates by Excel."""
    print("Generating B1_EXCEL_DATE_CORRUPTION.xlsx...")
    chems = get_industrial_chemicals(20)
    
    data = []
    for c in chems:
        cas = c['cas_number'] or ''
        # Corrupt some CAS numbers to dates
        if cas and random.random() < 0.5:
            cas = corrupt_cas_to_date(cas)
        
        data.append({
            'Chemical Name': c['name'],
            'CAS Number': cas,
            'Quantity': f"{random.randint(10, 500)} L",
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'B1_EXCEL_DATE_CORRUPTION.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_B2_mixed_encodings():
    """B2: Mixed character encodings (UTF-8, special chars)."""
    print("Generating B2_MIXED_ENCODINGS.csv...")
    
    # Use chemicals with special characters
    data = [
        {'Chemical Name': 'α-Naphthol', 'CAS': '90-15-3', 'Qty': '100 L'},
        {'Chemical Name': 'β-Carotene', 'CAS': '7235-40-7', 'Qty': '50 kg'},
        {'Chemical Name': 'Café (test)', 'CAS': '58-08-2', 'Qty': '25 L'},
        {'Chemical Name': 'Naïve compound', 'CAS': '50-00-0', 'Qty': '75 L'},
    ]
    
    # Add some normal chemicals
    chems = get_industrial_chemicals(16)
    for c in chems:
        data.append({
            'Chemical Name': c['name'],
            'CAS': c['cas_number'] or '',
            'Qty': f"{random.randint(10, 500)} L",
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'B2_MIXED_ENCODINGS.csv')
    df.to_csv(filepath, index=False, encoding='utf-8')
    print(f"  ✓ Created {filepath}")


def generate_B3_messy_quantities():
    """B3: Inconsistent quantity formatting."""
    print("Generating B3_MESSY_QUANTITIES.xlsx...")
    chems = get_industrial_chemicals(20)
    
    formats = [
        lambda: f"{random.randint(10, 500)} L",
        lambda: f"{random.randint(10, 500)}L",
        lambda: f"{random.randint(10, 500)}.{random.randint(0, 9)} kg",
        lambda: f"{random.randint(1, 10)},{random.randint(100, 999)} ml",
        lambda: f"~{random.randint(10, 100)} gal",
        lambda: f"{random.randint(100, 200)}-{random.randint(201, 300)} drums",
    ]
    
    data = []
    for c in chems:
        data.append({
            'Chemical Name': c['name'],
            'CAS': c['cas_number'] or '',
            'Quantity': random.choice(formats)(),
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'B3_MESSY_QUANTITIES.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_B4_empty_rows_and_columns():
    """B4: File with scattered empty rows and columns."""
    print("Generating B4_EMPTY_ROWS_AND_COLUMNS.xlsx...")
    chems = get_industrial_chemicals(15)
    
    filepath = os.path.join(OUTPUT_DIR, 'B4_EMPTY_ROWS_AND_COLUMNS.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 5 empty rows at top
    row_idx = 6
    
    # Headers with empty column in middle
    ws[f'A{row_idx}'] = 'Chemical Name'
    ws[f'B{row_idx}'] = 'CAS'
    # Column C empty
    ws[f'D{row_idx}'] = 'Quantity'
    ws[f'E{row_idx}'] = 'Unit'
    
    row_idx += 1
    
    # Data with scattered empty rows
    for i, c in enumerate(chems):
        ws[f'A{row_idx}'] = c['name']
        ws[f'B{row_idx}'] = c['cas_number'] or ''
        ws[f'D{row_idx}'] = random.randint(10, 500)
        ws[f'E{row_idx}'] = random.choice(['L', 'kg'])
        row_idx += 1
        
        # Insert empty row every 3 rows
        if (i + 1) % 3 == 0:
            row_idx += 1
    
    wb.save(filepath)
    print(f"  ✓ Created {filepath}")


def generate_B5_duplicate_column_names():
    """B5: File with duplicate column names."""
    print("Generating B5_DUPLICATE_COLUMN_NAMES.xlsx...")
    chems = get_industrial_chemicals(15)
    
    data = []
    for c in chems:
        data.append([
            c['name'],
            c['cas_number'] or '',
            random.randint(10, 500),  # Quantity 1
            random.randint(10, 500),  # Quantity 2 (duplicate)
            'Note 1',
            'Note 2',
            'Note 3',
        ])
    
    # Create DataFrame with duplicate column names
    df = pd.DataFrame(data, columns=[
        'Chemical Name', 'CAS', 'Quantity', 'Quantity', 'Notes', 'Notes', 'Notes'
    ])
    
    filepath = os.path.join(OUTPUT_DIR, 'B5_DUPLICATE_COLUMN_NAMES.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


# ═══════════════════════════════════════════════════════
#  Category C: Matching Engine Challenges
# ═══════════════════════════════════════════════════════

def generate_C1_missing_names():
    """C1: Rows with valid CAS but empty name, and vice versa."""
    print("Generating C1_MISSING_NAMES.xlsx...")
    chems = get_industrial_chemicals(20)
    
    data = []
    for i, c in enumerate(chems):
        if i < 10:
            # First 10: Valid CAS, empty name
            data.append({
                'Chemical Name': '',
                'CAS Number': c['cas_number'] or '',
                'Quantity': f"{random.randint(10, 500)} L",
            })
        else:
            # Last 10: Valid name, empty/invalid CAS
            data.append({
                'Chemical Name': c['name'],
                'CAS Number': '',
                'Quantity': f"{random.randint(10, 500)} L",
            })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'C1_MISSING_NAMES.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_C2_field_swap_detection():
    """C2: Field swaps - CAS in name column, name in CAS column."""
    print("Generating C2_FIELD_SWAP_DETECTION.xlsx...")
    chems = get_chemicals_with_formula(10)
    
    data = []
    for i, c in enumerate(chems):
        if i == 0:
            # Row 1: CAS in name column, name in CAS column
            data.append({
                'Chemical Name': c['cas_number'] or '',
                'CAS Number': c['name'],
                'Quantity': f"{random.randint(10, 500)} L",
            })
        elif i == 1:
            # Row 2: Name in formula column, formula in name column
            data.append({
                'Chemical Name': c.get('formula', ''),
                'CAS Number': c['cas_number'] or '',
                'Formula': c['name'],
            })
        else:
            # Normal rows
            data.append({
                'Chemical Name': c['name'],
                'CAS Number': c['cas_number'] or '',
                'Quantity': f"{random.randint(10, 500)} L",
            })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'C2_FIELD_SWAP_DETECTION.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_C3_fuzzy_name_matching():
    """C3: Names with typos and variations."""
    print("Generating C3_FUZZY_NAME_MATCHING.xlsx...")
    
    # Deliberate typos and variations
    data = [
        {'Chemical Name': 'ACETone', 'CAS': '', 'Qty': '100 L'},
        {'Chemical Name': 'Meth anol', 'CAS': '', 'Qty': '50 L'},
        {'Chemical Name': 'Ethyl Alcohol', 'CAS': '', 'Qty': '75 L'},
        {'Chemical Name': 'Calcuim Carbonate', 'CAS': '', 'Qty': '200 kg'},
        {'Chemical Name': 'Sodim Hydroxide', 'CAS': '', 'Qty': '150 L'},
        {'Chemical Name': 'Sulphuric Acid', 'CAS': '', 'Qty': '100 L'},
        {'Chemical Name': 'Hydro chloric Acid', 'CAS': '', 'Qty': '80 L'},
        {'Chemical Name': 'Amonia', 'CAS': '', 'Qty': '60 L'},
        {'Chemical Name': 'Benzne', 'CAS': '', 'Qty': '120 L'},
        {'Chemical Name': 'Tolune', 'CAS': '', 'Qty': '90 L'},
    ]
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'C3_FUZZY_NAME_MATCHING.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_C4_synonym_resolution():
    """C4: Trade names and synonyms."""
    print("Generating C4_SYNONYM_RESOLUTION.xlsx...")
    
    # Trade names that should resolve to canonical names
    data = [
        {'Chemical Name': 'Caustic Soda', 'CAS': '', 'Qty': '100 L'},
        {'Chemical Name': 'Muriatic Acid', 'CAS': '', 'Qty': '50 L'},
        {'Chemical Name': 'Wood Alcohol', 'CAS': '', 'Qty': '75 L'},
        {'Chemical Name': 'Rubbing Alcohol', 'CAS': '', 'Qty': '60 L'},
        {'Chemical Name': 'Spirits of Salt', 'CAS': '', 'Qty': '80 L'},
        {'Chemical Name': 'Oil of Vitriol', 'CAS': '', 'Qty': '90 L'},
        {'Chemical Name': 'Quicklime', 'CAS': '', 'Qty': '200 kg'},
        {'Chemical Name': 'Slaked Lime', 'CAS': '', 'Qty': '150 kg'},
        {'Chemical Name': 'Table Salt', 'CAS': '', 'Qty': '500 kg'},
        {'Chemical Name': 'Epsom Salt', 'CAS': '', 'Qty': '100 kg'},
    ]
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'C4_SYNONYM_RESOLUTION.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_C5_conflict_detection():
    """C5: Conflicting CAS/Name/UN combinations."""
    print("Generating C5_CONFLICT_DETECTION.xlsx...")
    
    # Deliberate conflicts
    data = [
        # Row 1: CAS for Acetone, Name for Methanol
        {'Chemical Name': 'METHANOL', 'CAS': '67-64-1', 'Qty': '100 L'},
        # Row 2: Name for Acetone, UN for Methanol
        {'Chemical Name': 'ACETONE', 'CAS': '', 'UN': '1230', 'Qty': '50 L'},
        # Row 3: Normal (no conflict)
        {'Chemical Name': 'ETHANOL', 'CAS': '64-17-5', 'Qty': '75 L'},
    ]
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'C5_CONFLICT_DETECTION.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


# ═══════════════════════════════════════════════════════
#  Category D: Real-World Production Scenarios
# ═══════════════════════════════════════════════════════

def generate_D1_petroleum_refinery():
    """D1: 50 rows of petroleum products."""
    print("Generating D1_PETROLEUM_REFINERY.xlsx...")
    chems = get_petroleum_chemicals(50)
    
    data = []
    for c in chems:
        data.append({
            'Product Name': c['name'],
            'CAS No.': c['cas_number'] or '',
            'Quantity': f"{random.randint(100, 10000)} L",
            'Location': random.choice(['Tank A', 'Tank B', 'Storage C']),
            'Supplier': random.choice(['Refinery X', 'Supplier Y', 'Distributor Z']),
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'D1_PETROLEUM_REFINERY.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_D2_industrial_chemicals():
    """D2: 50 rows of industrial chemicals."""
    print("Generating D2_INDUSTRIAL_CHEMICALS.xlsx...")
    chems = get_industrial_chemicals(50)
    
    data = []
    for c in chems:
        data.append({
            'Chemical': c['name'],
            'CAS': c['cas_number'] or '',
            'Stock': f"{random.randint(10, 1000)} {random.choice(['L', 'kg', 'gal'])}",
            'Location': f"Warehouse {random.choice(['A', 'B', 'C'])}-{random.randint(1, 20)}",
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'D2_INDUSTRIAL_CHEMICALS.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_D3_pharmaceuticals():
    """D3: 30 pharmaceutical ingredients (expected low coverage)."""
    print("Generating D3_PHARMACEUTICALS.xlsx...")
    
    # Pharma ingredients (most NOT in CAMEO)
    pharma_data = [
        {'Name': 'Paracetamol', 'CAS': '103-90-2'},
        {'Name': 'Ibuprofen', 'CAS': '15687-27-1'},
        {'Name': 'Aspirin', 'CAS': '50-78-2'},
        {'Name': 'Codeine', 'CAS': '76-57-3'},
        {'Name': 'Atropine', 'CAS': '51-55-8'},
        {'Name': 'Vitamin C', 'CAS': '50-81-7'},
        {'Name': 'Caffeine', 'CAS': '58-08-2'},
        {'Name': 'Nicotine', 'CAS': '54-11-5'},
        {'Name': 'Morphine', 'CAS': '57-27-2'},
        {'Name': 'Penicillin', 'CAS': '61-33-6'},
    ]
    
    # Repeat to get 30 rows
    data = []
    for i in range(30):
        p = pharma_data[i % len(pharma_data)]
        data.append({
            'Ingredient': p['Name'],
            'CAS Number': p['CAS'],
            'Batch': f"BATCH-{random.randint(1000, 9999)}",
            'Qty': f"{random.randint(1, 100)} kg",
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'D3_PHARMACEUTICALS.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


def generate_D4_multi_sheet_excel():
    """D4: Multi-sheet Excel with only one sheet containing chemical data."""
    print("Generating D4_MULTI_SHEET_EXCEL.xlsx...")
    chems = get_industrial_chemicals(20)
    
    filepath = os.path.join(OUTPUT_DIR, 'D4_MULTI_SHEET_EXCEL.xlsx')
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Sheet 1: Summary (no chemical data)
        summary = pd.DataFrame({
            'Report': ['Monthly Inventory Report'],
            'Date': [datetime.now().strftime('%Y-%m-%d')],
            'Total Items': [len(chems)],
        })
        summary.to_excel(writer, sheet_name='Summary', index=False)
        
        # Sheet 2: Inventory (actual chemical data)
        inventory = []
        for c in chems:
            inventory.append({
                'Chemical Name': c['name'],
                'CAS': c['cas_number'] or '',
                'Quantity': f"{random.randint(10, 500)} L",
            })
        pd.DataFrame(inventory).to_excel(writer, sheet_name='Inventory', index=False)
        
        # Sheet 3: Archived (empty)
        pd.DataFrame().to_excel(writer, sheet_name='Archived', index=False)
        
        # Sheet 4: Notes (text only)
        notes = pd.DataFrame({
            'Notes': ['Check expiry dates', 'Reorder acetone', 'Update MSDS']
        })
        notes.to_excel(writer, sheet_name='Notes', index=False)
    
    print(f"  ✓ Created {filepath}")


def generate_D5_large_file():
    """D5: 500 rows of mixed chemicals."""
    print("Generating D5_LARGE_FILE.xlsx...")
    
    # Mix of petroleum, industrial, and chemicals with UN/formula
    petroleum = get_petroleum_chemicals(200)
    industrial = get_industrial_chemicals(200)
    with_un = get_chemicals_with_un(100)
    
    all_chems = petroleum + industrial + with_un
    random.shuffle(all_chems)
    all_chems = all_chems[:500]
    
    data = []
    for c in all_chems:
        data.append({
            'Chemical Name': c['name'],
            'CAS Number': c['cas_number'] or '',
            'Quantity': f"{random.randint(10, 5000)} {random.choice(['L', 'kg', 'gal', 'drums'])}",
            'Location': f"{random.choice(['A', 'B', 'C', 'D'])}-{random.randint(1, 50)}",
            'Supplier': f"Supplier {random.randint(1, 10)}",
            'Date Received': (datetime.now() - timedelta(days=random.randint(1, 365))).strftime('%Y-%m-%d'),
        })
    
    df = pd.DataFrame(data)
    filepath = os.path.join(OUTPUT_DIR, 'D5_LARGE_FILE.xlsx')
    df.to_excel(filepath, index=False)
    print(f"  ✓ Created {filepath}")


# ═══════════════════════════════════════════════════════
#  Main Generator
# ═══════════════════════════════════════════════════════

def main():
    """Generate all 20 test files."""
    print("=" * 60)
    print("SAFEWARE ETL COMPREHENSIVE STRESS TEST FILE GENERATOR")
    print("=" * 60)
    print()
    
    # Category A: Column Detection
    print("Category A: Column Detection Challenges")
    generate_A1_no_headers()
    generate_A2_foreign_headers()
    generate_A3_abbreviated_headers()
    generate_A4_swapped_columns()
    generate_A5_merged_headers()
    print()
    
    # Category B: Data Quality
    print("Category B: Data Quality Challenges")
    generate_B1_excel_date_corruption()
    generate_B2_mixed_encodings()
    generate_B3_messy_quantities()
    generate_B4_empty_rows_and_columns()
    generate_B5_duplicate_column_names()
    print()
    
    # Category C: Matching Engine
    print("Category C: Matching Engine Challenges")
    generate_C1_missing_names()
    generate_C2_field_swap_detection()
    generate_C3_fuzzy_name_matching()
    generate_C4_synonym_resolution()
    generate_C5_conflict_detection()
    print()
    
    # Category D: Real-World Scenarios
    print("Category D: Real-World Production Scenarios")
    generate_D1_petroleum_refinery()
    generate_D2_industrial_chemicals()
    generate_D3_pharmaceuticals()
    generate_D4_multi_sheet_excel()
    generate_D5_large_file()
    print()
    
    print("=" * 60)
    print("✓ ALL 20 TEST FILES GENERATED SUCCESSFULLY")
    print(f"✓ Output directory: {OUTPUT_DIR}")
    print("=" * 60)


if __name__ == '__main__':
    main()
